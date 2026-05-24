from flask import Flask, request, redirect, url_for, flash, render_template_string, session, send_file
from wtforms import StringField, IntegerField, SelectField, SelectMultipleField, DateField, RadioField, HiddenField
from wtforms.widgets import CheckboxInput, ListWidget
from wtforms.validators import DataRequired, Optional, NumberRange, ValidationError
from flask_wtf import FlaskForm
from flask_wtf.csrf import CSRFProtect
from datetime import date
import os
import json
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side

app = Flask(__name__)
app.secret_key = 'your-secret-key-change-in-production'
csrf = CSRFProtect(app)

# -------------------- البحث عن ملف الإكسل --------------------
EXCEL_XLSFORM = None
for candidate in ['اختباري القراءة والحساب.xlsx', 'eee.xlsx']:
    if os.path.exists(candidate):
        EXCEL_XLSFORM = candidate
        break
if EXCEL_XLSFORM is None:
    print("⚠️ تحذير: لم يتم العثور على ملف الإكسل. سيتم استخدام بيانات افتراضية.")
    GOV_CHOICES = [('1','الحديدة'),('2','الضالع'),('3','شبوه'),('4','عدن'),('5','لحج')]
    DIST_BY_GOV = {}
    SCHOOL_BY_DIST = {}
else:
    print(f"✅ تم استخدام ملف الإكسل: {EXCEL_XLSFORM}")

# -------------------- إعدادات قاعدة بيانات المدخلين --------------------
USERS_FILE = "data/users.xlsx"
DATA_DIR = "data"
EXCEL_FILE = os.path.join(DATA_DIR, "submissions.xlsx")
os.makedirs(DATA_DIR, exist_ok=True)

def init_users():
    if not os.path.exists(USERS_FILE):
        df = pd.DataFrame(columns=["username", "password", "fullname"])
        df.loc[0] = ["admin", "admin123", "مدير النظام"]
        df.to_excel(USERS_FILE, index=False, engine='openpyxl')

def get_all_users():
    if not os.path.exists(USERS_FILE): init_users()
    return pd.read_excel(USERS_FILE, engine='openpyxl').to_dict(orient='records')

def add_user(username, password, fullname):
    init_users()
    df = pd.read_excel(USERS_FILE, engine='openpyxl')
    if username in df['username'].values:
        return False
    df.loc[len(df)] = [username, password, fullname]
    df.to_excel(USERS_FILE, index=False, engine='openpyxl')
    return True

def authenticate_user(username, password):
    users = get_all_users()
    for u in users:
        if u['username'] == username and u['password'] == password:
            return u
    return None

# -------------------- إعدادات حفظ بيانات التقييم --------------------
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        columns = ["submission_id", "timestamp", "ip_address", "enumerator", "a0",
                   "p1_1", "p1_2", "p1_3", "p1_4", "school_code",
                   "researcher_name", "researcher_id", "researcher_gender",
                   "student_name1", "student_name2", "student_name3", "student_name4",
                   "student_gender", "student_age", "student_grade", "student_stream", "student_residence", "student_disability",
                   "hand_name1", "hand_name2", "hand_name3", "hand_name4",
                   "final_reading_level", "final_math_level"]
        for i in range(1,6): columns.append(f'R3_{i}')
        columns.append('L1_3')
        columns.append('R4_fluent')
        columns.append('L1_4')
        for i in range(1,4): columns.append(f'R5_q{i}')
        columns.append('L1_5')
        for i in range(1,6): columns.append(f'R2_{i}')
        for i in range(1,11): columns.append(f'R1_{i}')
        for i in range(1,4): columns.append(f'M3_{i}')
        for i in range(1,4): columns.append(f'M4_{i}')
        for i in range(1,3): columns.append(f'M5_{i}')
        for i in range(1,6): columns.append(f'M2_{i}')
        for i in range(1,6): columns.append(f'M1_{i}')
        df = pd.DataFrame(columns=columns)
        df.to_excel(EXCEL_FILE, index=False, engine='openpyxl')
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        header_fill = PatternFill(start_color="1a4d8c", end_color="1a4d8c", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        wb.save(EXCEL_FILE)

def format_excel():
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        header_fill = PatternFill(start_color="1a4d8c", end_color="1a4d8c", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        wb.save(EXCEL_FILE)

def get_label(value, list_name):
    if not EXCEL_XLSFORM:
        return value
    try:
        df = pd.read_excel(EXCEL_XLSFORM, sheet_name='choices', dtype=str)
        df = df.fillna('')
        row = df[(df['list_name'] == list_name) & (df['name'] == str(value))]
        if not row.empty:
            label = row.iloc[0].get('label::Arabic', '')
            if label:
                return label
        return value
    except:
        return value

def save_submission(data_dict, ip_address, enumerator):
    init_excel()
    df = pd.read_excel(EXCEL_FILE, engine='openpyxl')
    submission_id = len(df) + 1
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    new_row = {"submission_id": submission_id, "timestamp": timestamp, "ip_address": ip_address, "enumerator": enumerator}
    
    for col in df.columns:
        if col in ["submission_id", "timestamp", "ip_address", "enumerator"]:
            continue
        value = data_dict.get(col, '')
        if col == 'a0':
            new_row['a0'] = 'نعم' if value == '1' else 'لا' if value == '2' else ''
        elif col == 'p1_2':
            new_row['p1_2'] = get_label(value, 'governorate')
        elif col == 'p1_3':
            new_row['p1_3'] = get_label(value, 'district')
        elif col == 'p1_4':
            new_row['p1_4'] = get_label(value, 'school')
        elif col == 'student_gender':
            new_row['student_gender'] = 'ذكر' if value == '1' else 'أنثى' if value == '2' else ''
        elif col == 'researcher_gender':
            new_row['researcher_gender'] = 'ذكر' if value == '1' else 'أنثى' if value == '2' else ''
        elif col == 'student_grade':
            new_row['student_grade'] = get_label(value, 'grade')
        elif col == 'student_stream':
            new_row['student_stream'] = get_label(value, 'stream')
        elif col == 'student_residence':
            new_row['student_residence'] = 'من أبناء المنطقة' if value == '1' else 'نازح' if value == '2' else ''
        elif col == 'disabilities':
            if isinstance(value, str):
                disabilities = value.split(',')
            else:
                disabilities = value if isinstance(value, list) else []
            disability_labels = []
            for d in disabilities:
                lbl = get_label(d, 'Yes_No_ch')
                disability_labels.append(lbl)
            new_row['student_disability'] = ','.join(disability_labels)
        elif col in ['R3_1','R3_2','R3_3','R3_4','R3_5','R4_fluent','R5_q1','R5_q2','R5_q3','R2_1','R2_2','R2_3','R2_4','R2_5','R1_1','R1_2','R1_3','R1_4','R1_5','R1_6','R1_7','R1_8','R1_9','R1_10','M3_1','M3_2','M3_3','M4_1','M4_2','M4_3','M5_1','M5_2','M2_1','M2_2','M2_3','M2_4','M2_5','M1_1','M1_2','M1_3','M1_4','M1_5']:
            new_row[col] = 'نعم' if value == '1' else 'لا' if value == '2' else value
        elif col in ['L1_3','L1_4','L1_5']:
            if value and value.isdigit():
                new_row[col] = get_label(value, col)
            else:
                new_row[col] = value
        else:
            new_row[col] = value
    
    final_r = data_dict.get('final_reading_level', '')
    final_m = data_dict.get('final_math_level', '')
    if not final_r:
        final_r = compute_reading_level(data_dict)
    if not final_m:
        final_m = compute_math_level(data_dict)
    new_row['final_reading_level'] = final_r
    new_row['final_math_level'] = final_m
    
    for col in df.columns:
        if col not in new_row:
            new_row[col] = ""
    df.loc[len(df)] = new_row
    df.to_excel(EXCEL_FILE, index=False, engine='openpyxl')
    format_excel()
    print(f"✅ تم حفظ الإدخال {submission_id} - القراءة: {final_r}, الحساب: {final_m}")
    return submission_id

def get_all_submissions():
    init_excel()
    return pd.read_excel(EXCEL_FILE, engine='openpyxl').to_dict(orient='records')

def get_enumerator_stats():
    df = pd.read_excel(EXCEL_FILE, engine='openpyxl')
    if 'enumerator' not in df.columns:
        return {}
    return df.groupby('enumerator').size().to_dict()

# -------------------- قراءة ملف XLSForm --------------------
def load_choices():
    if not EXCEL_XLSFORM:
        return {}
    df = pd.read_excel(EXCEL_XLSFORM, sheet_name='choices', dtype=str)
    df = df.fillna('')
    choices_dict = {}
    for _, row in df.iterrows():
        list_name = row['list_name']
        if list_name not in choices_dict:
            choices_dict[list_name] = []
        label = row.get('label::Arabic', '')
        if not label:
            label = row.get('label', '')
        choices_dict[list_name].append({
            'name': str(row['name']),
            'label': label
        })
    return choices_dict

def load_governorates_districts_schools():
    if not EXCEL_XLSFORM:
        return [], {}, {}
    df = pd.read_excel(EXCEL_XLSFORM, sheet_name='choices', dtype=str)
    df = df.fillna('')
    govs = df[df['list_name'] == 'governorate']
    gov_choices = [(row['name'], row['label::Arabic']) for _, row in govs.iterrows() if row.get('name') and row.get('label::Arabic')]
    districts_df = df[df['list_name'] == 'district']
    dist_by_gov = {}
    for _, row in districts_df.iterrows():
        gov = str(row.get('governorate', '')).strip()
        name = str(row.get('name', '')).strip()
        label = row.get('label::Arabic', '')
        if gov and name and label:
            dist_by_gov.setdefault(gov, []).append({'value': name, 'label': label})
    schools_df = df[df['list_name'] == 'school']
    school_by_dist = {}
    for _, row in schools_df.iterrows():
        dist = str(row.get('district', '')).strip()
        name = str(row.get('name', '')).strip()
        label = row.get('label::Arabic', '')
        if dist and name and label:
            school_by_dist.setdefault(dist, []).append({'value': name, 'label': label})
    return gov_choices, dist_by_gov, school_by_dist

GOV_CHOICES, DIST_BY_GOV, SCHOOL_BY_DIST = load_governorates_districts_schools()
if not GOV_CHOICES:
    GOV_CHOICES = [('1','الحديدة'),('2','الضالع'),('3','شبوه'),('4','عدن'),('5','لحج')]
if not DIST_BY_GOV:
    DIST_BY_GOV = {g[0]: [{'value': '0', 'label': 'لا توجد مديريات'}] for g in GOV_CHOICES}
if not SCHOOL_BY_DIST:
    SCHOOL_BY_DIST = {}

# -------------------- دوال حساب المستوى --------------------
def compute_reading_level(data):
    R3_correct = sum(1 for i in range(1,6) if data.get(f'R3_{i}') == '1')
    R3_attempt = sum(1 for i in range(1,6) if data.get(f'R3_{i}') == '2')
    if R3_correct >= 4:
        if data.get('R4_fluent') == '1':
            R5_correct = sum(1 for i in range(1,4) if data.get(f'R5_q{i}') == '1')
            return 5 if R5_correct >= 2 else 4
        else:
            return 4
    elif R3_attempt >= 2:
        R2_correct = sum(1 for i in range(1,6) if data.get(f'R2_{i}') == '1')
        if R2_correct >= 4:
            return 2
        else:
            R2_attempt = sum(1 for i in range(1,6) if data.get(f'R2_{i}') == '2')
            if R2_attempt >= 2:
                R1_correct = sum(1 for i in range(1,11) if data.get(f'R1_{i}') == '1')
                if R1_correct >= 8:
                    return 1
                else:
                    R1_attempt = sum(1 for i in range(1,11) if data.get(f'R1_{i}') == '2')
                    return 0 if R1_attempt >= 3 else 1
            else:
                return 2
    else:
        return 3

def compute_math_level(data):
    M3_correct = sum(1 for i in range(1,4) if data.get(f'M3_{i}') == '1')
    M3_attempt = sum(1 for i in range(1,4) if data.get(f'M3_{i}') == '2')
    if M3_correct >= 2:
        M4_correct = sum(1 for i in range(1,4) if data.get(f'M4_{i}') == '1')
        if M4_correct >= 2:
            M5_correct = sum(1 for i in range(1,3) if data.get(f'M5_{i}') == '1')
            return 5 if M5_correct == 2 else 4
        else:
            return 4
    elif M3_attempt >= 2:
        M2_correct = sum(1 for i in range(1,6) if data.get(f'M2_{i}') == '1')
        if M2_correct >= 4:
            return 2
        else:
            M2_attempt = sum(1 for i in range(1,6) if data.get(f'M2_{i}') == '2')
            if M2_attempt >= 2:
                M1_correct = sum(1 for i in range(1,6) if data.get(f'M1_{i}') == '1')
                if M1_correct >= 4:
                    return 1
                else:
                    M1_attempt = sum(1 for i in range(1,6) if data.get(f'M1_{i}') == '2')
                    return 0 if M1_attempt >= 2 else 1
            else:
                return 2
    else:
        return 3

class MultiCheckboxField(SelectMultipleField):
    widget = ListWidget(prefix_label=False)
    option_widget = CheckboxInput()

# -------------------- نموذج الاستبيان --------------------
class AssessmentForm(FlaskForm):
    start = HiddenField()
    end = HiddenField()
    today = HiddenField(default=date.today().isoformat())
    deviceid = HiddenField(default="web")
    a0 = RadioField('هل يمكننا البدء في التقييم؟', choices=[('1','نعم'),('2','لا')], validators=[DataRequired()])
    p1_1 = DateField('تاريخ المقابلة (اليوم/الشهر/السنة)', validators=[Optional()], format='%Y-%m-%d')
    p1_2 = SelectField('المحافظة', choices=GOV_CHOICES, validators=[Optional()])
    p1_3 = SelectField('المديرية', choices=[('', '-- اختر --')], validators=[Optional()])
    p1_4 = SelectField('المدرسة', choices=[('', '-- اختر --')], validators=[Optional()])
    researcher_name = StringField('اسم الباحث (المعلم)/ الباحثة (المعلمة)', validators=[Optional()])
    researcher_id = StringField('رقم البطاقة الشخصية للباحث / الباحثة', validators=[Optional()])
    researcher_gender = SelectField('النوع الاجتماعي للباحث / الباحثة؟', choices=[('1','ذكر'),('2','أنثى')], validators=[Optional()])
    student_name1 = StringField('الاسم الأول', validators=[Optional()])
    student_name2 = StringField('الاسم الثاني', validators=[Optional()])
    student_name3 = StringField('الاسم الثالث', validators=[Optional()])
    student_name4 = StringField('الاسم الرابع', validators=[Optional()])
    student_gender = SelectField('النوع الاجتماعي؟', choices=[('1','ذكر'),('2','أنثى')], validators=[Optional()])
    student_age = IntegerField('عمر التلميذ/ التلميذة بالسنوات؟', validators=[Optional(), NumberRange(min=5, max=15, message="العمر يجب أن يكون بين 5 و 15 سنة")])
    student_grade = SelectField('المستوى الدراسي العام للتلميذ/ للتلميذة؟', choices=[('1','صف ثاني'),('2','صف ثالث'),('3','صف رابع')], validators=[Optional()])
    student_stream = SelectField('الشعبة الدراسية للتلميذ/ للتلميذة؟', choices=[('1','شعبة أ'),('2','شعبة ب'),('3','شعبة ج'),('4','شعبة د'),('5','شعبة ه'),('6','شعبة و'),('7','شعبة ز')], validators=[Optional()])
    student_residence = SelectField('حالة الإقامة؟', choices=[('1','من أبناء المنطقة'),('2','نازح')], validators=[Optional()])
    disability_choices = load_choices().get('Yes_No_ch', [])
    disabilities = MultiCheckboxField('هل يعاني التلميذ/ التلميذة من اي اعاقة؟', choices=[(c['name'], c['label']) for c in disability_choices], validators=[Optional()])
    hand_name1 = StringField('الاسم الأول (بخط اليد)', validators=[Optional()])
    hand_name2 = StringField('الاسم الثاني (بخط اليد)', validators=[Optional()])
    hand_name3 = StringField('الاسم الثالث (بخط اليد)', validators=[Optional()])
    hand_name4 = StringField('الاسم الرابع (بخط اليد)', validators=[Optional()])
    for i in range(1,6):
        vars()[f'R3_{i}'] = RadioField(f'هل قرأ الجملة {i}؟', choices=[('1','نعم'),('2','لا')], validators=[Optional()])
    L1_3 = SelectField('مستوى الإتقان؟', choices=[], validators=[Optional()])
    R4_fluent = RadioField('هل قرأ القصة كاملة بطلاقة؟', choices=[('1','نعم'),('2','لا')], validators=[Optional()])
    L1_4 = SelectField('مستوى الأتقان؟', choices=[], validators=[Optional()])
    for i in range(1,4):
        vars()[f'R5_q{i}'] = RadioField(f'السؤال {i}', choices=[('1','نعم'),('2','لا')], validators=[Optional()])
    L1_5 = SelectField('مستوى الأتقان؟', choices=[], validators=[Optional()])
    for i in range(1,6):
        vars()[f'R2_{i}'] = RadioField(f'قرأ الكلمة {i}؟', choices=[('1','نعم'),('2','لا')], validators=[Optional()])
    for i in range(1,11):
        vars()[f'R1_{i}'] = RadioField(f'قرأ الحرف {i}؟', choices=[('1','نعم'),('2','لا')], validators=[Optional()])
    final_reading_level = HiddenField()
    for i in range(1,4):
        vars()[f'M3_{i}'] = RadioField(f'مسألة جمع {i}', choices=[('1','نعم'),('2','لا')], validators=[Optional()])
    for i in range(1,4):
        vars()[f'M4_{i}'] = RadioField(f'مسألة طرح {i}', choices=[('1','نعم'),('2','لا')], validators=[Optional()])
    for i in range(1,3):
        vars()[f'M5_{i}'] = RadioField(f'مسألة قسمة {i}', choices=[('1','نعم'),('2','لا')], validators=[Optional()])
    for i in range(1,6):
        vars()[f'M2_{i}'] = RadioField(f'العدد {i}', choices=[('1','نعم'),('2','لا')], validators=[Optional()])
    for i in range(1,6):
        vars()[f'M1_{i}'] = RadioField(f'الرقم {i}', choices=[('1','نعم'),('2','لا')], validators=[Optional()])
    final_math_level = HiddenField()

# -------------------- قالب HTML (نسخة كاملة تعمل) --------------------
FORM_HTML = '''
<!DOCTYPE html>
<html dir="rtl" lang="ar">
<head><meta charset="UTF-8"><title>اختبار القراءة والحساب</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
<link href="https://fonts.googleapis.com/css2?family=Cairo:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
    body{background: linear-gradient(135deg, #e0eafc 0%, #cfdef3 100%); font-family: 'Cairo', sans-serif;}
    .form-container{max-width:1000px;margin:30px auto;background:#ffffff;padding:35px;border-radius:30px;box-shadow:0 20px 40px rgba(0,0,0,0.1);}
    .section-title{background:linear-gradient(95deg,#1a4d8c,#2a6fb0);color:white;padding:14px 20px;border-radius:50px;margin:30px 0 25px;font-weight:600;}
    .level-title{background:#f0f7ff;border-right:6px solid #1a4d8c;padding:10px 18px;margin:20px 0 15px;font-weight:600;color:#1a4d8c;border-radius:12px;}
    .btn-submit{background:linear-gradient(95deg,#1a4d8c,#2a6fb0);color:white;font-size:1.2rem;padding:14px;border-radius:50px;width:100%;margin-top:25px;border:none;}
    .btn-submit:hover{transform:translateY(-2px);box-shadow:0 8px 20px rgba(0,0,0,0.15);}
    .hidden-group{display:none;}
    .card-radio{border:1px solid #e2e8f0;border-radius:20px;padding:12px 18px;margin-bottom:12px;background:#fafcff; display: flex; align-items: center; justify-content: space-between; flex-wrap: wrap;}
    .card-radio .radio-group { display: flex; gap: 15px; align-items: center; }
    .card-radio label { font-weight: 500; margin: 0; }
    .card-radio .form-check { margin-left: 15px; }
    .checkbox-group{display:flex;flex-wrap:wrap;gap:18px;margin-top:8px;}
    .level-display{background:#eef4ff;padding:12px 18px;border-radius:30px;font-weight:600;margin-top:15px;text-align:center;color:#1a4d8c; border: 2px solid #1a4d8c;}
    .footer-copyright{text-align:center;margin-top:35px;padding-top:15px;border-top:1px solid #dee2e6;color:#6c757d;font-size:0.85rem;}
    .alert-floating{position:fixed;top:20px;left:50%;transform:translateX(-50%);z-index:9999;min-width:300px;text-align:center;box-shadow:0 4px 12px rgba(0,0,0,0.15);}
    .level-group { border: 1px solid #cfdef3; border-radius: 20px; padding: 15px; margin-bottom: 25px; background-color: #ffffff; }
    .vertical-radio { display: flex; flex-direction: column; gap: 5px; margin-top: 5px; }
    .vertical-radio .form-check { margin-bottom: 5px; }
    .consent-text { background: #f8f9fa; padding: 15px; border-radius: 15px; margin-bottom: 15px; }
</style>
</head>
<body>
<div class="container form-container">
    <h2 class="text-center" style="color:#1a4d8c; font-weight:700;">تقييم حالة التعلم والتعليم السنوية (الأيسر) - اختبــاري القراءة والحساب - مشروع استدامة التعليم والتعلم</h2>
    <p class="text-center text-muted">نظام متابعة الأداء – يرجى تعبئة البيانات بدقة</p>
    <hr>
    {% with messages = get_flashed_messages() %}
        {% if messages %}
            <div class="alert alert-success alert-floating">{{ messages[0] }}</div>
        {% endif %}
    {% endwith %}
    <form method="POST" id="mainForm">
        {{ form.csrf_token }}
        
        <div class="consent-text">
            <h5>إفادة الموافقة الشفوية</h5>
            <p>مرحباً، كيف حالكم؟ أنا أسمي _________________. نحن نقوم بإجراء تقييم لمستوى التعلم و التعليم. نحن نقدر كثيراً مشاركتك في هذا التقييم.<br>
            أنا هنا مع زملائي، نحن حقا نقدر مشاركتك في هذه النقاشات، وهذه الأسئلة ستأخذ قرابة 20 دقيقة. المعلومات التي ستذكرونها ستبقى آمنة وسرية. سيتم الاحتفاظ بأي معلومات تُقدم لغرض تطوير التعليم. المشاركة في هذا الاستطلاع طواعية وليس إجباراً.<br>
            إذا لم تحصل على الموافقة، اشكر التلميذ/ة وانتقل للتلميذ/ة التالي.</p>
        </div>
        
        <div class="alert alert-info">
            <div class="vertical-radio">
                {{ form.a0.label(class="form-label fw-bold") }}
                {% for subfield in form.a0 %}
                    <div class="form-check">
                        {{ subfield(class="form-check-input") }}
                        <label class="form-check-label">{{ subfield.label.text }}</label>
                    </div>
                {% endfor %}
            </div>
        </div>
        
        <div id="secA" class="hidden-group">
            <div class="section-title">📋 القسم الأول – التعريف</div>
            <div class="mb-2">{{ form.p1_1.label(class="form-label fw-bold") }} {{ form.p1_1(class="form-control") }}</div>
            <div class="mb-2">{{ form.p1_2.label(class="form-label fw-bold") }} {{ form.p1_2(class="form-select", id="gov") }}</div>
            <div class="mb-2">{{ form.p1_3.label(class="form-label fw-bold") }} {{ form.p1_3(class="form-select", id="dist") }}</div>
            <div class="mb-2">{{ form.p1_4.label(class="form-label fw-bold") }} {{ form.p1_4(class="form-select", id="sch") }}</div>
            <div id="schoolCodeNote" class="alert alert-secondary mt-2" style="display:none;">🏫 كود المدرسة: <span id="schoolCodeValue"></span></div>
            <div class="mb-2">{{ form.researcher_name.label(class="form-label fw-bold") }} {{ form.researcher_name(class="form-control", id="researcher_name") }}</div>
            <div class="mb-2">{{ form.researcher_id.label(class="form-label fw-bold") }} {{ form.researcher_id(class="form-control", id="researcher_id") }}</div>
            <div class="mb-2">{{ form.researcher_gender.label(class="form-label fw-bold") }} {{ form.researcher_gender(class="form-select", id="researcher_gender") }}</div>
        </div>
        
        <div id="secB" class="hidden-group">
            <div class="section-title">👤 القسم الثاني - المعلومات الاساسية</div>
            <div class="mb-2">{{ form.student_name1.label(class="form-label fw-bold") }} {{ form.student_name1(class="form-control", id="student_name1") }}</div>
            <div class="mb-2">{{ form.student_name2.label(class="form-label fw-bold") }} {{ form.student_name2(class="form-control", id="student_name2") }}</div>
            <div class="mb-2">{{ form.student_name3.label(class="form-label fw-bold") }} {{ form.student_name3(class="form-control", id="student_name3") }}</div>
            <div class="mb-2">{{ form.student_name4.label(class="form-label fw-bold") }} {{ form.student_name4(class="form-control", id="student_name4") }}</div>
            <div class="mb-2">{{ form.student_gender.label(class="form-label fw-bold") }} {{ form.student_gender(class="form-select", id="student_gender") }}</div>
            <div class="mb-2">{{ form.student_age.label(class="form-label fw-bold") }} {{ form.student_age(class="form-control", type="number", min=5, max=15, id="student_age") }}</div>
            <div class="mb-2">{{ form.student_grade.label(class="form-label fw-bold") }} {{ form.student_grade(class="form-select", id="student_grade") }}</div>
            <div class="mb-2">{{ form.student_stream.label(class="form-label fw-bold") }} {{ form.student_stream(class="form-select", id="student_stream") }}</div>
            <div class="mb-2">{{ form.student_residence.label(class="form-label fw-bold") }} {{ form.student_residence(class="form-select", id="student_residence") }}</div>
            <div class="mb-2">{{ form.disabilities.label(class="form-label fw-bold") }} <div class="checkbox-group" id="disabilityGroup">{{ form.disabilities() }}</div></div>
            <div class="level-title">✍️ أطلب من التلميذ/ التلميذة كتابة اسمه / اسمها الكل بخط يده/ يدها هنا</div>
            <div class="mb-2">{{ form.hand_name1.label(class="form-label") }} {{ form.hand_name1(class="form-control", id="hand_name1") }} <span id="warning1" class="text-danger small"></span></div>
            <div class="mb-2">{{ form.hand_name2.label(class="form-label") }} {{ form.hand_name2(class="form-control", id="hand_name2") }} <span id="warning2" class="text-danger small"></span></div>
            <div class="mb-2">{{ form.hand_name3.label(class="form-label") }} {{ form.hand_name3(class="form-control", id="hand_name3") }} <span id="warning3" class="text-danger small"></span></div>
            <div class="mb-2">{{ form.hand_name4.label(class="form-label") }} {{ form.hand_name4(class="form-control", id="hand_name4") }} <span id="warning4" class="text-danger small"></span>
        </div>
        
        <div id="readingSec" class="hidden-group">
            <div class="section-title">📖 أولاً: اختبار القراءة</div>
            <div id="level3Reading" class="level-group">
                <div class="level-title">المستوى الثالث: جمل – مقطع</div>
                {% for i in range(1,6) %}
                <div class="card-radio">
                    <label>{{ form['R3_'~i].label }}</label>
                    <div class="radio-group">
                        {% for subfield in form['R3_'~i] %}
                            <div class="form-check form-check-inline">
                                {{ subfield(class="form-check-input") }}
                                <label class="form-check-label">{{ subfield.label.text }}</label>
                            </div>
                        {% endfor %}
                    </div>
                </div>
                {% endfor %}
                <div class="mt-2" id="L1_3_container">{{ form.L1_3.label }} {{ form.L1_3(class="form-select", id="L1_3") }}</div>
            </div>
            <div id="level4Reading" class="level-group hidden-group">
                <div class="level-title">المستوى الرابع: قصة</div>
                <div class="card-radio">
                    <label>{{ form.R4_fluent.label }}</label>
                    <div class="radio-group">
                        {% for subfield in form.R4_fluent %}
                            <div class="form-check form-check-inline">
                                {{ subfield(class="form-check-input") }}
                                <label class="form-check-label">{{ subfield.label.text }}</label>
                            </div>
                        {% endfor %}
                    </div>
                </div>
                <div class="mt-2" id="L1_4_container">{{ form.L1_4.label }} {{ form.L1_4(class="form-select", id="L1_4") }}</div>
            </div>
            <div id="level5Reading" class="level-group hidden-group">
                <div class="level-title">المستوى الخامس: الفهم</div>
                {% for i in range(1,4) %}
                <div class="card-radio">
                    <label>{{ form['R5_q'~i].label }}</label>
                    <div class="radio-group">
                        {% for subfield in form['R5_q'~i] %}
                            <div class="form-check form-check-inline">
                                {{ subfield(class="form-check-input") }}
                                <label class="form-check-label">{{ subfield.label.text }}</label>
                            </div>
                        {% endfor %}
                    </div>
                </div>
                {% endfor %}
                <div class="mt-2" id="L1_5_container">{{ form.L1_5.label }} {{ form.L1_5(class="form-select", id="L1_5") }}</div>
            </div>
            <div id="level2Reading" class="level-group hidden-group">
                <div class="level-title">المستوى الثاني: كلمات</div>
                {% for i in range(1,6) %}
                <div class="card-radio">
                    <label>{{ form['R2_'~i].label }}</label>
                    <div class="radio-group">
                        {% for subfield in form['R2_'~i] %}
                            <div class="form-check form-check-inline">
                                {{ subfield(class="form-check-input") }}
                                <label class="form-check-label">{{ subfield.label.text }}</label>
                            </div>
                        {% endfor %}
                    </div>
                </div>
                {% endfor %}
            </div>
            <div id="level1Reading" class="level-group hidden-group">
                <div class="level-title">المستوى الأول: حروف</div>
                {% for i in range(1,11) %}
                <div class="card-radio">
                    <label>{{ form['R1_'~i].label }}</label>
                    <div class="radio-group">
                        {% for subfield in form['R1_'~i] %}
                            <div class="form-check form-check-inline">
                                {{ subfield(class="form-check-input") }}
                                <label class="form-check-label">{{ subfield.label.text }}</label>
                            </div>
                        {% endfor %}
                    </div>
                </div>
                {% endfor %}
            </div>
            <div class="level-display">🏆 مستوى القراءة الحالي: <span id="readingLevel" class="fw-bold">-</span></div>
        </div>
        
        <div id="mathSec" class="hidden-group" style="display:none;">
            <div class="section-title">🧮 ثانياً: اختبار الحساب</div>
            <div id="level3Math" class="level-group">
                <div class="level-title">المستوى الثالث: الجمع مع الحمل</div>
                {% for i in range(1,4) %}
                <div class="card-radio">
                    <label>{{ form['M3_'~i].label }}</label>
                    <div class="radio-group">
                        {% for subfield in form['M3_'~i] %}
                            <div class="form-check form-check-inline">
                                {{ subfield(class="form-check-input") }}
                                <label class="form-check-label">{{ subfield.label.text }}</label>
                            </div>
                        {% endfor %}
                    </div>
                </div>
                {% endfor %}
            </div>
            <div id="level4Math" class="level-group hidden-group">
                <div class="level-title">المستوى الرابع: الطرح مع الاستلاف</div>
                {% for i in range(1,4) %}
                <div class="card-radio">
                    <label>{{ form['M4_'~i].label }}</label>
                    <div class="radio-group">
                        {% for subfield in form['M4_'~i] %}
                            <div class="form-check form-check-inline">
                                {{ subfield(class="form-check-input") }}
                                <label class="form-check-label">{{ subfield.label.text }}</label>
                            </div>
                        {% endfor %}
                    </div>
                </div>
                {% endfor %}
            </div>
            <div id="level5Math" class="level-group hidden-group">
                <div class="level-title">المستوى الخامس: القسمة مع الحمل</div>
                {% for i in range(1,3) %}
                <div class="card-radio">
                    <label>{{ form['M5_'~i].label }}</label>
                    <div class="radio-group">
                        {% for subfield in form['M5_'~i] %}
                            <div class="form-check form-check-inline">
                                {{ subfield(class="form-check-input") }}
                                <label class="form-check-label">{{ subfield.label.text }}</label>
                            </div>
                        {% endfor %}
                    </div>
                </div>
                {% endfor %}
            </div>
            <div id="level2Math" class="level-group hidden-group">
                <div class="level-title">المستوى الثاني: الأعداد من ١٠ إلى ٩٩</div>
                {% for i in range(1,6) %}
                <div class="card-radio">
                    <label>{{ form['M2_'~i].label }}</label>
                    <div class="radio-group">
                        {% for subfield in form['M2_'~i] %}
                            <div class="form-check form-check-inline">
                                {{ subfield(class="form-check-input") }}
                                <label class="form-check-label">{{ subfield.label.text }}</label>
                            </div>
                        {% endfor %}
                    </div>
                </div>
                {% endfor %}
            </div>
            <div id="level1Math" class="level-group hidden-group">
                <div class="level-title">المستوى الأول: الأعداد من ١ إلى ٩</div>
                {% for i in range(1,6) %}
                <div class="card-radio">
                    <label>{{ form['M1_'~i].label }}</label>
                    <div class="radio-group">
                        {% for subfield in form['M1_'~i] %}
                            <div class="form-check form-check-inline">
                                {{ subfield(class="form-check-input") }}
                                <label class="form-check-label">{{ subfield.label.text }}</label>
                            </div>
                        {% endfor %}
                    </div>
                </div>
                {% endfor %}
            </div>
            <div class="level-display">🏆 مستوى الحساب الحالي: <span id="mathLevel" class="fw-bold">-</span></div>
        </div>
        
        <button type="submit" class="btn-submit">💾 إرسال البيانات</button>
        <input type="hidden" name="final_reading_level" id="final_reading_level">
        <input type="hidden" name="final_math_level" id="final_math_level">
        <input type="hidden" name="school_code" id="school_code_hidden">
    </form>
    <div class="footer-copyright">© abdulbaqi alkamali - 2026. جميع الحقوق محفوظة</div>
</div>
<script>
    const distByGov = {{ dist_by_gov | safe }};
    const schoolByDist = {{ school_by_dist | safe }};
    const govSel = document.getElementById('gov');
    const distSel = document.getElementById('dist');
    const schSel = document.getElementById('sch');
    const schoolCodeNote = document.getElementById('schoolCodeNote');
    const schoolCodeValue = document.getElementById('schoolCodeValue');
    const schoolCodeHidden = document.getElementById('school_code_hidden');
    const researcherName = document.getElementById('researcher_name');
    const researcherId = document.getElementById('researcher_id');
    const researcherGender = document.getElementById('researcher_gender');
    
    function saveResearcherDataForSchool(schoolCode) {
        if (!schoolCode) return;
        const data = { name: researcherName.value, id: researcherId.value, gender: researcherGender.value };
        localStorage.setItem(`researcher_${schoolCode}`, JSON.stringify(data));
    }
    function loadResearcherDataForSchool(schoolCode) {
        if (!schoolCode) {
            researcherName.value = ''; researcherId.value = ''; researcherGender.value = '';
            return;
        }
        const saved = localStorage.getItem(`researcher_${schoolCode}`);
        if (saved) {
            const data = JSON.parse(saved);
            researcherName.value = data.name || '';
            researcherId.value = data.id || '';
            researcherGender.value = data.gender || '';
        } else {
            researcherName.value = ''; researcherId.value = ''; researcherGender.value = '';
        }
    }
    
    function updateDistricts() {
        let gov = govSel.value;
        distSel.innerHTML = '<option value="">-- اختر --</option>';
        schSel.innerHTML = '<option value="">-- اختر --</option>';
        schoolCodeNote.style.display = 'none';
        if (gov && distByGov[gov]) {
            distByGov[gov].forEach(d => {
                let opt = document.createElement('option');
                opt.value = d.value;
                opt.text = d.label;
                distSel.appendChild(opt);
            });
        }
        researcherName.value = ''; researcherId.value = ''; researcherGender.value = '';
    }
    function updateSchools() {
        let dist = distSel.value;
        schSel.innerHTML = '<option value="">-- اختر --</option>';
        schoolCodeNote.style.display = 'none';
        if (dist && schoolByDist[dist]) {
            schoolByDist[dist].forEach(s => {
                let opt = document.createElement('option');
                opt.value = s.value;
                opt.text = s.label;
                schSel.appendChild(opt);
            });
        }
    }
    function showSchoolCode() {
        let selectedOption = schSel.options[schSel.selectedIndex];
        if (selectedOption && selectedOption.value) {
            let schoolCode = selectedOption.value;
            schoolCodeValue.innerText = schoolCode;
            schoolCodeNote.style.display = 'block';
            schoolCodeHidden.value = schoolCode;
            loadResearcherDataForSchool(schoolCode);
        } else {
            schoolCodeNote.style.display = 'none';
            schoolCodeHidden.value = '';
            loadResearcherDataForSchool('');
        }
    }
    researcherName.addEventListener('change', () => { if (schoolCodeHidden.value) saveResearcherDataForSchool(schoolCodeHidden.value); });
    researcherId.addEventListener('change', () => { if (schoolCodeHidden.value) saveResearcherDataForSchool(schoolCodeHidden.value); });
    researcherGender.addEventListener('change', () => { if (schoolCodeHidden.value) saveResearcherDataForSchool(schoolCodeHidden.value); });
    
    govSel.addEventListener('change', updateDistricts);
    distSel.addEventListener('change', updateSchools);
    schSel.addEventListener('change', showSchoolCode);
    
    const a0Radios = document.querySelectorAll('input[name="a0"]');
    const secA = document.getElementById('secA');
    const secB = document.getElementById('secB');
    const readingSec = document.getElementById('readingSec');
    const mathSec = document.getElementById('mathSec');
    function toggleSections() {
        let checked = document.querySelector('input[name="a0"]:checked');
        let show = checked && checked.value === '1';
        secA.style.display = show ? 'block' : 'none';
        secB.style.display = show ? 'block' : 'none';
        readingSec.style.display = show ? 'block' : 'none';
    }
    a0Radios.forEach(r => r.addEventListener('change', toggleSections));
    
    const disabilityCheckboxes = document.querySelectorAll('#disabilityGroup input[type="checkbox"]');
    const noDisabilityCheckbox = document.querySelector('#disabilityGroup input[value="0"]');
    function handleDisabilityChange() {
        if (noDisabilityCheckbox && noDisabilityCheckbox.checked) {
            disabilityCheckboxes.forEach(cb => { if (cb !== noDisabilityCheckbox) cb.disabled = true; });
        } else {
            disabilityCheckboxes.forEach(cb => cb.disabled = false);
        }
    }
    if (noDisabilityCheckbox) {
        noDisabilityCheckbox.addEventListener('change', handleDisabilityChange);
        disabilityCheckboxes.forEach(cb => { if (cb !== noDisabilityCheckbox) cb.addEventListener('change', handleDisabilityChange); });
    }
    
    const studentName1 = document.getElementById('student_name1');
    const studentName2 = document.getElementById('student_name2');
    const studentName3 = document.getElementById('student_name3');
    const studentName4 = document.getElementById('student_name4');
    const handName1 = document.getElementById('hand_name1');
    const handName2 = document.getElementById('hand_name2');
    const handName3 = document.getElementById('hand_name3');
    const handName4 = document.getElementById('hand_name4');
    const warn1 = document.getElementById('warning1');
    const warn2 = document.getElementById('warning2');
    const warn3 = document.getElementById('warning3');
    const warn4 = document.getElementById('warning4');
    function compareNames(orig, hand, warnSpan) {
        if (hand.value.trim() !== "" && orig.value.trim() !== hand.value.trim()) {
            warnSpan.innerText = "⚠️ يختلف عن الاسم المسجل أعلاه";
        } else {
            warnSpan.innerText = "";
        }
    }
    if (studentName1 && handName1) handName1.addEventListener('input', () => compareNames(studentName1, handName1, warn1));
    if (studentName2 && handName2) handName2.addEventListener('input', () => compareNames(studentName2, handName2, warn2));
    if (studentName3 && handName3) handName3.addEventListener('input', () => compareNames(studentName3, handName3, warn3));
    if (studentName4 && handName4) handName4.addEventListener('input', () => compareNames(studentName4, handName4, warn4));
    
    const L1_3 = document.getElementById('L1_3');
    const L1_4 = document.getElementById('L1_4');
    const L1_5 = document.getElementById('L1_5');
    function setOptions(select, opts) {
        select.innerHTML = '';
        opts.forEach(opt => {
            let o = document.createElement('option');
            o.value = opt.value;
            o.text = opt.text;
            select.appendChild(o);
        });
    }
    function updateL1_3Options() {
        let correct = 0;
        for(let i=1;i<=5;i++) {
            let r = document.querySelector(`input[name="R3_${i}"]:checked`);
            if(r && r.value === '1') correct++;
        }
        if(correct >= 4) {
            setOptions(L1_3, [{'value':'3','text':'بطلاقة'},{'value':'2','text':'بعض الصعوبة'}]);
        } else {
            setOptions(L1_3, [{'value':'1','text':'الصعوبة كبيرة'},{'value':'0','text':'لا مجال - لا يستطيع'},{'value':'999','text':'لم يتم التقييم'}]);
        }
        L1_3.value = '';
    }
    function updateL1_4Options() {
        let fluent = document.querySelector('input[name="R4_fluent"]:checked');
        if(fluent && fluent.value === '1') {
            setOptions(L1_4, [{'value':'3','text':'بطلاقة'}]);
        } else {
            setOptions(L1_4, [{'value':'2','text':'بعض الصعوبة'},{'value':'1','text':'الصعوبة كبيرة'},{'value':'0','text':'لا مجال - لا يستطيع'},{'value':'999','text':'لم يتم التقييم'}]);
        }
        L1_4.value = '';
    }
    function updateL1_5Options() {
        let correct = 0;
        for(let i=1;i<=3;i++) {
            let r = document.querySelector(`input[name="R5_q${i}"]:checked`);
            if(r && r.value === '1') correct++;
        }
        if(correct >= 2) {
            setOptions(L1_5, [{'value':'3','text':'بطلاقة'},{'value':'2','text':'بعض الصعوبة'}]);
        } else {
            setOptions(L1_5, [{'value':'1','text':'الصعوبة كبيرة'},{'value':'0','text':'لا مجال - لا يستطيع'},{'value':'999','text':'لم يتم التقييم'}]);
        }
        L1_5.value = '';
    }
    
    function getReadingData() {
        let data = {};
        for(let i=1;i<=5;i++) {
            let e = document.querySelector(`input[name="R3_${i}"]:checked`);
            if(e) data[`R3_${i}`] = e.value;
        }
        let f = document.querySelector('input[name="R4_fluent"]:checked');
        if(f) data['R4_fluent'] = f.value;
        for(let i=1;i<=3;i++) {
            let e = document.querySelector(`input[name="R5_q${i}"]:checked`);
            if(e) data[`R5_q${i}`] = e.value;
        }
        for(let i=1;i<=5;i++) {
            let e = document.querySelector(`input[name="R2_${i}"]:checked`);
            if(e) data[`R2_${i}`] = e.value;
        }
        for(let i=1;i<=10;i++) {
            let e = document.querySelector(`input[name="R1_${i}"]:checked`);
            if(e) data[`R1_${i}`] = e.value;
        }
        return data;
    }
    function computeReadingLevel(d) {
        let c3=0, a3=0;
        for(let i=1;i<=5;i++) {
            let v = d[`R3_${i}`];
            if(v === '1') c3++;
            else if(v === '2') a3++;
        }
        if(c3 >= 4) {
            if(d['R4_fluent'] === '1') {
                let c5=0;
                for(let i=1;i<=3;i++) if(d[`R5_q${i}`] === '1') c5++;
                return c5 >= 2 ? 5 : 4;
            } else { return 4; }
        } else if(a3 >= 2) {
            let c2=0, a2=0;
            for(let i=1;i<=5;i++) {
                let v = d[`R2_${i}`];
                if(v === '1') c2++;
                else if(v === '2') a2++;
            }
            if(c2 >= 4) return 2;
            else if(a2 >= 2) {
                let c1=0;
                for(let i=1;i<=10;i++) if(d[`R1_${i}`] === '1') c1++;
                if(c1 >= 8) return 1;
                else {
                    let a1=0;
                    for(let i=1;i<=10;i++) if(d[`R1_${i}`] === '2') a1++;
                    return a1 >= 3 ? 0 : 1;
                }
            } else return 2;
        } else return 3;
    }
    function updateReadingDisplay() {
        let lvl = computeReadingLevel(getReadingData());
        document.getElementById('readingLevel').innerText = lvl;
        document.getElementById('final_reading_level').value = lvl;
        if (lvl !== '' && lvl !== null && lvl !== '-') {
            mathSec.style.display = 'block';
        } else {
            mathSec.style.display = 'none';
        }
    }
    function updateReadingVisibility() {
        let c3 = 0, a3 = 0;
        let allR3Answered = true;
        for(let i=1;i<=5;i++) {
            let e = document.querySelector(`input[name="R3_${i}"]:checked`);
            if(!e) { allR3Answered = false; break; }
            if(e.value === '1') c3++;
            else if(e.value === '2') a3++;
        }
        const level4 = document.getElementById('level4Reading');
        const level5 = document.getElementById('level5Reading');
        const level2 = document.getElementById('level2Reading');
        const level1 = document.getElementById('level1Reading');
        
        if(allR3Answered) {
            if(c3 >= 4) {
                if(L1_3.value !== '') {
                    level4.classList.remove('hidden-group');
                    level2.classList.add('hidden-group');
                    level5.classList.add('hidden-group');
                    level1.classList.add('hidden-group');
                    let fluent = document.querySelector('input[name="R4_fluent"]:checked');
                    if(fluent && fluent.value === '1' && L1_4.value !== '') {
                        level5.classList.remove('hidden-group');
                    } else {
                        level5.classList.add('hidden-group');
                    }
                } else {
                    level4.classList.add('hidden-group');
                }
            } else if(a3 >= 2) {
                if(L1_3.value !== '') {
                    level4.classList.add('hidden-group');
                    level2.classList.remove('hidden-group');
                    level5.classList.add('hidden-group');
                    level1.classList.add('hidden-group');
                    let c2 = 0, a2 = 0, allR2Answered = true;
                    for(let i=1;i<=5;i++) {
                        let e = document.querySelector(`input[name="R2_${i}"]:checked`);
                        if(!e) { allR2Answered = false; break; }
                        if(e.value === '1') c2++;
                        else if(e.value === '2') a2++;
                    }
                    if(allR2Answered && c2 >= 4) {
                        level1.classList.add('hidden-group');
                    } else if(allR2Answered && a2 >= 2) {
                        level1.classList.remove('hidden-group');
                    } else {
                        level1.classList.add('hidden-group');
                    }
                } else {
                    level2.classList.add('hidden-group');
                }
            } else {
                level4.classList.add('hidden-group');
                level2.classList.add('hidden-group');
                level5.classList.add('hidden-group');
                level1.classList.add('hidden-group');
            }
        } else {
            level4.classList.add('hidden-group');
            level2.classList.add('hidden-group');
            level5.classList.add('hidden-group');
            level1.classList.add('hidden-group');
        }
    }
    
    function getMathData() {
        let data = {};
        for(let i=1;i<=3;i++) { let e = document.querySelector(`input[name="M3_${i}"]:checked`); if(e) data[`M3_${i}`] = e.value; }
        for(let i=1;i<=3;i++) { let e = document.querySelector(`input[name="M4_${i}"]:checked`); if(e) data[`M4_${i}`] = e.value; }
        for(let i=1;i<=2;i++) { let e = document.querySelector(`input[name="M5_${i}"]:checked`); if(e) data[`M5_${i}`] = e.value; }
        for(let i=1;i<=5;i++) { let e = document.querySelector(`input[name="M2_${i}"]:checked`); if(e) data[`M2_${i}`] = e.value; }
        for(let i=1;i<=5;i++) { let e = document.querySelector(`input[name="M1_${i}"]:checked`); if(e) data[`M1_${i}`] = e.value; }
        return data;
    }
    function computeMathLevel(d) {
        let c3=0; for(let i=1;i<=3;i++) if(d[`M3_${i}`]==='1') c3++;
        if(c3 >= 2) {
            let c4=0; for(let i=1;i<=3;i++) if(d[`M4_${i}`]==='1') c4++;
            if(c4 >= 2) {
                let c5=0; for(let i=1;i<=2;i++) if(d[`M5_${i}`]==='1') c5++;
                return c5 === 2 ? 5 : 4;
            } else return 4;
        } else {
            let a3=0; for(let i=1;i<=3;i++) if(d[`M3_${i}`]==='2') a3++;
            if(a3 >= 2) {
                let c2=0; for(let i=1;i<=5;i++) if(d[`M2_${i}`]==='1') c2++;
                if(c2 >= 4) return 2;
                else {
                    let a2=0; for(let i=1;i<=5;i++) if(d[`M2_${i}`]==='2') a2++;
                    if(a2 >= 2) {
                        let c1=0; for(let i=1;i<=5;i++) if(d[`M1_${i}`]==='1') c1++;
                        if(c1 >= 4) return 1;
                        else {
                            let a1=0; for(let i=1;i<=5;i++) if(d[`M1_${i}`]==='2') a1++;
                            return a1 >= 2 ? 0 : 1;
                        }
                    } else return 2;
                }
            } else return 3;
        }
    }
    function updateMathDisplay() {
        let lvl = computeMathLevel(getMathData());
        document.getElementById('mathLevel').innerText = lvl;
        document.getElementById('final_math_level').value = lvl;
    }
    function updateMathVisibility() {
        let c3=0,a3=0;
        let allM3Answered = true;
        for(let i=1;i<=3;i++) {
            let e = document.querySelector(`input[name="M3_${i}"]:checked`);
            if(!e) { allM3Answered = false; break; }
            if(e.value==='1') c3++;
            else if(e.value==='2') a3++;
        }
        const l4=document.getElementById('level4Math');
        const l5=document.getElementById('level5Math');
        const l2=document.getElementById('level2Math');
        const l1=document.getElementById('level1Math');
        if(allM3Answered) {
            if(c3 >= 2) {
                l4.classList.remove('hidden-group');
                l2.classList.add('hidden-group');
                l5.classList.add('hidden-group');
                l1.classList.add('hidden-group');
                let allM4Answered = true, c4=0;
                for(let i=1;i<=3;i++) {
                    let e = document.querySelector(`input[name="M4_${i}"]:checked`);
                    if(!e) { allM4Answered = false; break; }
                    if(e.value==='1') c4++;
                }
                if(c4 >= 2 && allM4Answered) l5.classList.remove('hidden-group');
            } else if(a3 >= 2) {
                l4.classList.add('hidden-group');
                l2.classList.remove('hidden-group');
                l5.classList.add('hidden-group');
                l1.classList.add('hidden-group');
                let allM2Answered = true, c2=0;
                for(let i=1;i<=5;i++) {
                    let e = document.querySelector(`input[name="M2_${i}"]:checked`);
                    if(!e) { allM2Answered = false; break; }
                    if(e.value==='1') c2++;
                }
                if(c2 >= 4 && allM2Answered) l1.classList.add('hidden-group');
                else if(allM2Answered) l1.classList.remove('hidden-group');
            } else {
                l4.classList.add('hidden-group');
                l2.classList.add('hidden-group');
                l5.classList.add('hidden-group');
                l1.classList.add('hidden-group');
            }
        } else {
            l4.classList.add('hidden-group');
            l2.classList.add('hidden-group');
            l5.classList.add('hidden-group');
            l1.classList.add('hidden-group');
        }
    }
    
    for(let i=1;i<=5;i++) {
        document.querySelectorAll(`input[name="R3_${i}"]`).forEach(r => r.addEventListener('change', () => {
            updateL1_3Options();
            updateReadingDisplay();
            updateReadingVisibility();
        }));
        document.querySelectorAll(`input[name="R2_${i}"]`).forEach(r => r.addEventListener('change', () => {
            updateReadingDisplay();
            updateReadingVisibility();
        }));
    }
    document.querySelectorAll('input[name="R4_fluent"]').forEach(r => r.addEventListener('change', () => {
        updateL1_4Options();
        updateReadingDisplay();
        updateReadingVisibility();
    }));
    for(let i=1;i<=3;i++) {
        document.querySelectorAll(`input[name="R5_q${i}"]`).forEach(r => r.addEventListener('change', () => {
            updateL1_5Options();
            updateReadingDisplay();
            updateReadingVisibility();
        }));
    }
    for(let i=1;i<=10;i++) {
        document.querySelectorAll(`input[name="R1_${i}"]`).forEach(r => r.addEventListener('change', updateReadingDisplay));
    }
    L1_3.addEventListener('change', () => {
        updateReadingVisibility();
        updateReadingDisplay();
    });
    L1_4.addEventListener('change', () => {
        updateReadingVisibility();
        updateReadingDisplay();
    });
    L1_5.addEventListener('change', () => {
        updateReadingVisibility();
        updateReadingDisplay();
    });
    
    for(let i=1;i<=3;i++) {
        document.querySelectorAll(`input[name="M3_${i}"]`).forEach(r => r.addEventListener('change', () => {
            updateMathDisplay();
            updateMathVisibility();
        }));
        document.querySelectorAll(`input[name="M4_${i}"]`).forEach(r => r.addEventListener('change', () => {
            updateMathDisplay();
            updateMathVisibility();
        }));
    }
    for(let i=1;i<=2;i++) document.querySelectorAll(`input[name="M5_${i}"]`).forEach(r => r.addEventListener('change', updateMathDisplay));
    for(let i=1;i<=5;i++) document.querySelectorAll(`input[name="M2_${i}"]`).forEach(r => r.addEventListener('change', () => {
        updateMathDisplay();
        updateMathVisibility();
    }));
    for(let i=1;i<=5;i++) document.querySelectorAll(`input[name="M1_${i}"]`).forEach(r => r.addEventListener('change', updateMathDisplay));
    
    window.onload = function() {
        toggleSections();
        updateDistricts();
        updateSchools();
        updateL1_3Options();
        updateL1_4Options();
        updateL1_5Options();
        updateReadingDisplay();
        updateMathDisplay();
        updateReadingVisibility();
        updateMathVisibility();
        if(govSel.value) updateDistricts();
        if(distSel.value) updateSchools();
        handleDisabilityChange();
        setTimeout(() => {
            document.querySelectorAll('.alert-floating').forEach(a => a.style.display = 'none');
        }, 5000);
        document.querySelectorAll('input[name="a0"]').forEach(r => r.checked = false);
        mathSec.style.display = 'none';
    };
</script>
</body>
</html>
'''

# -------------------- القوالب الأخرى --------------------
LOGIN_HTML = '''
<!DOCTYPE html>
<html dir="rtl" lang="ar">
<head><meta charset="UTF-8"><title>دخول المدخل</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
</head>
<body style="background:#f0f2f5"><div class="container" style="max-width:400px;margin-top:100px;background:#fff;padding:30px;border-radius:20px">
<h2 class="text-center">تسجيل دخول المدخل</h2>
{% with messages = get_flashed_messages() %}{% if messages %}<div class="alert alert-danger">{{ messages[0] }}</div>{% endif %}{% endwith %}
<form method="POST"><input type="hidden" name="csrf_token" value="{{ csrf_token() }}">
<div class="mb-3"><label>اسم المستخدم</label> <input type="text" name="username" class="form-control" required></div>
<div class="mb-3"><label>كلمة المرور</label> <input type="password" name="password" class="form-control" required></div>
<button type="submit" class="btn btn-primary w-100">دخول</button>
</form><div class="text-center mt-3">© abdulbaqi alkamali - 2026</div>
</div></body></html>
'''

ADMIN_HTML = '''
<!DOCTYPE html>
<html dir="rtl" lang="ar">
<head><meta charset="UTF-8"><title>لوحة المدير</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0-beta3/css/all.min.css">
<style>
    body{background: #f0f2f5; font-family: 'Cairo', sans-serif;}
    .dashboard-card{border-radius: 20px; transition: transform 0.2s, box-shadow 0.2s; border: none;}
    .dashboard-card:hover{transform: translateY(-5px); box-shadow: 0 15px 30px rgba(0,0,0,0.1);}
    .btn-download{background: linear-gradient(135deg, #28a745, #20c997); border: none;}
    .btn-download:hover{background: linear-gradient(135deg, #218838, #1aa179);}
    .btn-upload{background: linear-gradient(135deg, #ffc107, #fd7e14); border: none; color: #000;}
    .btn-upload:hover{background: linear-gradient(135deg, #e0a800, #e06e0a);}
</style>
</head>
<body>
<div class="container mt-5">
    <div class="text-center mb-5">
        <h2 style="color:#1a4d8c;"><i class="fas fa-tachometer-alt"></i> لوحة تحكم المدير</h2>
        <p class="text-muted">إدارة المدخلين ومتابعة البيانات</p>
    </div>
    <div class="row">
        <div class="col-md-4 mb-4">
            <div class="card dashboard-card text-center p-3 shadow-sm">
                <div class="card-body">
                    <i class="fas fa-user-plus fa-3x text-primary"></i>
                    <h5 class="mt-2">إضافة مدخل جديد</h5>
                    <a href="/admin/users" class="btn btn-primary mt-2 rounded-pill px-4">إضافة</a>
                </div>
            </div>
        </div>
        <div class="col-md-4 mb-4">
            <div class="card dashboard-card text-center p-3 shadow-sm">
                <div class="card-body">
                    <i class="fas fa-upload fa-3x text-warning"></i>
                    <h5 class="mt-2">رفع ملف مدخلين (Excel)</h5>
                    <form action="/admin/upload-users" method="POST" enctype="multipart/form-data" class="mt-2">
                        <input type="file" name="users_file" accept=".xlsx, .xls" class="form-control form-control-sm mb-2" required>
                        <button type="submit" class="btn btn-upload rounded-pill px-3"><i class="fas fa-cloud-upload-alt"></i> رفع</button>
                    </form>
                </div>
            </div>
        </div>
        <div class="col-md-4 mb-4">
            <div class="card dashboard-card text-center p-3 shadow-sm">
                <div class="card-body">
                    <i class="fas fa-download fa-3x text-success"></i>
                    <h5 class="mt-2">تحميل بيانات الإدخالات</h5>
                    <a href="/admin/download-submissions" class="btn btn-download mt-2 rounded-pill px-4"><i class="fas fa-file-excel"></i> تحميل</a>
                </div>
            </div>
        </div>
        <div class="col-md-4 mb-4">
            <div class="card dashboard-card text-center p-3 shadow-sm">
                <div class="card-body">
                    <i class="fas fa-table fa-3x text-info"></i>
                    <h5 class="mt-2">عرض جميع الإدخالات</h5>
                    <a href="/admin/submissions" class="btn btn-info mt-2 rounded-pill px-4 text-white">عرض</a>
                </div>
            </div>
        </div>
        <div class="col-md-4 mb-4">
            <div class="card dashboard-card text-center p-3 shadow-sm">
                <div class="card-body">
                    <i class="fas fa-chart-bar fa-3x text-secondary"></i>
                    <h5 class="mt-2">إحصائيات المدخلين</h5>
                    <a href="/admin/stats" class="btn btn-secondary mt-2 rounded-pill px-4">عرض</a>
                </div>
            </div>
        </div>
    </div>
    <div class="text-center mt-4">
        <a href="/logout" class="btn btn-danger rounded-pill px-5"><i class="fas fa-sign-out-alt"></i> تسجيل خروج</a>
    </div>
    <div class="text-center mt-4 text-muted">© abdulbaqi alkamali - 2026</div>
</div>
</body>
</html>
'''

UPLOAD_USERS_HTML = '''
<!DOCTYPE html>
<html dir="rtl" lang="ar">
<head><meta charset="UTF-8"><title>رفع المدخلين</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
</head>
<body><div class="container mt-5"><div class="card p-4"><h2 class="text-center">رفع ملف مدخلين</h2>
<p>الملف يجب أن يحتوي على الأعمدة: <strong>username, password, fullname</strong></p>
<form method="POST" enctype="multipart/form-data"><input type="file" name="users_file" class="form-control" required><button type="submit" class="btn btn-primary mt-3">رفع</button></form>
<a href="/admin" class="btn btn-secondary mt-3">رجوع</a></div></div></body></html>
'''

USERS_HTML = '''
<!DOCTYPE html>
<html dir="rtl" lang="ar">
<head><meta charset="UTF-8"><title>إدارة المدخلين</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
</head>
<body><div class="container mt-5"><div class="card p-4"><h2 class="text-center">إضافة مدخل بيانات</h2>
{% with messages = get_flashed_messages() %}{% if messages %}<div class="alert alert-success">{{ messages[0] }}</div>{% endif %}{% endwith %}
<form method="POST"><input type="hidden" name="csrf_token" value="{{ csrf_token() }}">
<div class="mb-3"><label>اسم المستخدم</label> <input type="text" name="username" class="form-control" required></div>
<div class="mb-3"><label>كلمة المرور</label> <input type="text" name="password" class="form-control" required></div>
<div class="mb-3"><label>الاسم الكامل</label> <input type="text" name="fullname" class="form-control" required></div>
<button type="submit" class="btn btn-primary w-100">إضافة</button>
</form><hr><h4>المدخلون الحاليون</h4><ul class="list-group">{% for u in users %}<li class="list-group-item">{{ u.fullname }} ({{ u.username }})</li>{% endfor %}</ul>
<a href="/admin" class="btn btn-secondary mt-3">رجوع</a><div class="text-center mt-3">© abdulbaqi alkamali - 2026</div>
</div></div></body></html>
'''

SUBMISSIONS_HTML = '''
<!DOCTYPE html>
<html dir="rtl" lang="ar">
<head><meta charset="UTF-8"><title>جميع الإدخالات</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
</head>
<body><div class="container mt-4"><div class="card p-4"><h2 class="text-center">جميع بيانات الاختبارات</h2><a href="/admin" class="btn btn-secondary">رجوع</a>
<div class="table-responsive"><table class="table table-striped"><thead class="table-dark"><tr><th>#</th><th>التاريخ</th><th>المدخل</th><th>المحافظة</th><th>المديرية</th><th>المدرسة</th><th>الاسم الأول</th><th>العمر</th><th>مستوى القراءة</th><th>مستوى الحساب</th></tr></thead>
<tbody>{% for s in submissions %}
    <tr>
        <td>{{ s.submission_id }}</td>
        <td>{{ s.timestamp }}</td>
        <td>{{ s.enumerator }}</td>
        <td>{{ s.p1_2 }}</td>
        <td>{{ s.p1_3 }}</td>
        <td>{{ s.p1_4 }}</td>
        <td>{{ s.student_name1 }}</td>
        <td>{{ s.student_age }}</td>
        <td>{{ s.final_reading_level }}</td>
        <td>{{ s.final_math_level }}</td>
    </tr>
{% endfor %}</tbody></table></div><div class="text-center mt-3">© abdulbaqi alkamali - 2026</div></div></div></body></html>
'''

STATS_HTML = '''
<!DOCTYPE html>
<html dir="rtl" lang="ar">
<head><meta charset="UTF-8"><title>إحصائيات المدخلين</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
</head>
<body><div class="container mt-5"><div class="card p-4"><h2 class="text-center">عدد الحالات لكل مدخل</h2><a href="/admin" class="btn btn-secondary">رجوع</a>
<table class="table"><thead><tr><th>المدخل</th><th>عدد الحالات</th></tr></thead>
<tbody>{% for name, count in stats.items() %}
    <tr><td class="tg-0lax">{{ name }}</td>
    <td>{{ count }}</td>
    </tr>
{% endfor %}</tbody>
</table><div class="text-center mt-3">© abdulbaqi alkamali - 2026</div></div></div></body></html>
'''

# -------------------- Routes --------------------
@app.route('/')
def home():
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = authenticate_user(username, password)
        if user:
            session['username'] = user['username']
            session['fullname'] = user['fullname']
            if username == 'admin':
                return redirect(url_for('admin_dashboard'))
            else:
                return redirect(url_for('assessment'))
        else:
            flash('اسم المستخدم أو كلمة المرور غير صحيحة')
    return render_template_string(LOGIN_HTML)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/admin')
def admin_dashboard():
    if session.get('username') != 'admin':
        return redirect(url_for('login'))
    return render_template_string(ADMIN_HTML)

@app.route('/admin/users', methods=['GET', 'POST'])
def manage_users():
    if session.get('username') != 'admin':
        return redirect(url_for('login'))
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        fullname = request.form['fullname']
        if add_user(username, password, fullname):
            flash('تمت الإضافة بنجاح')
        else:
            flash('اسم المستخدم موجود مسبقاً')
        return redirect(url_for('manage_users'))
    users = get_all_users()
    return render_template_string(USERS_HTML, users=users)

@app.route('/admin/submissions')
def view_submissions():
    if session.get('username') != 'admin':
        return redirect(url_for('login'))
    submissions = get_all_submissions()
    return render_template_string(SUBMISSIONS_HTML, submissions=submissions)

@app.route('/admin/stats')
def view_stats():
    if session.get('username') != 'admin':
        return redirect(url_for('login'))
    stats = get_enumerator_stats()
    return render_template_string(STATS_HTML, stats=stats)

@app.route('/admin/upload-users', methods=['GET', 'POST'])
def upload_users():
    if session.get('username') != 'admin':
        return redirect(url_for('login'))
    if request.method == 'POST':
        file = request.files.get('users_file')
        if not file or file.filename == '':
            flash('❌ يرجى اختيار ملف Excel صحيح.', 'danger')
            return redirect(url_for('upload_users'))
        try:
            df = pd.read_excel(file, engine='openpyxl')
            required_cols = ['username', 'password', 'fullname']
            if not all(col in df.columns for col in required_cols):
                flash('⚠️ الملف يجب أن يحتوي على الأعمدة: username, password, fullname', 'danger')
                return redirect(url_for('upload_users'))
            added = 0
            skipped = 0
            for _, row in df.iterrows():
                username = str(row['username']).strip()
                password = str(row['password']).strip()
                fullname = str(row['fullname']).strip()
                if username and password and fullname:
                    if add_user(username, password, fullname):
                        added += 1
                    else:
                        skipped += 1
                else:
                    skipped += 1
            flash(f'✅ تمت إضافة {added} مدخل جديد. تم تخطي {skipped} (موجود أو ناقص).', 'success')
        except Exception as e:
            flash(f'❌ خطأ في قراءة الملف: {e}', 'danger')
        return redirect(url_for('admin_dashboard'))
    return render_template_string(UPLOAD_USERS_HTML)

@app.route('/admin/download-submissions')
def download_submissions():
    file_path = os.path.join(DATA_DIR, "submissions.xlsx")
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True, download_name="submissions.xlsx")
    else:
        flash("⚠️ لا يوجد ملف بيانات بعد. قم بإدخال بعض البيانات أولاً.", "warning")
        return redirect(url_for('admin_dashboard'))

@app.route('/assessment', methods=['GET', 'POST'])
def assessment():
    if 'username' not in session or session.get('username') == 'admin':
        return redirect(url_for('login'))
    if not GOV_CHOICES:
        flash('⚠️ لم يتم تحميل بيانات المحافظات والمديريات والمدارس. يرجى التأكد من وجود ملف الإكسل (اختباري القراءة والحساب.xlsx أو eee.xlsx) في نفس المجلد.')
        return redirect(url_for('login'))
    form = AssessmentForm()
    if request.method == 'POST':
        a0_value = request.form.get('a0')
        if a0_value == '2':
            data = {'a0': '2'}
            save_submission(data, request.remote_addr, session.get('fullname', ''))
            flash('✅ تم تسجيل رفض التلميذ للاختبار.')
            return redirect(url_for('assessment'))
        elif a0_value == '1':
            if not request.form.get('L1_3'):
                flash('❌ يرجى اختيار مستوى الإتقان للمستوى الثالث.')
                return redirect(url_for('assessment'))
            if request.form.get('R4_fluent') and not request.form.get('L1_4'):
                flash('❌ يرجى اختيار مستوى الإتقان للمستوى الرابع.')
                return redirect(url_for('assessment'))
            r5_q1 = request.form.get('R5_q1')
            r5_q2 = request.form.get('R5_q2')
            r5_q3 = request.form.get('R5_q3')
            if (r5_q1 or r5_q2 or r5_q3) and not request.form.get('L1_5'):
                flash('❌ يرجى اختيار مستوى الإتقان للمستوى الخامس.')
                return redirect(url_for('assessment'))
            
            data = request.form.to_dict()
            if 'disabilities' in request.form:
                data['disabilities'] = ','.join(request.form.getlist('disabilities'))
            else:
                data['disabilities'] = ''
            if 'final_reading_level' not in data or not data['final_reading_level']:
                data['final_reading_level'] = compute_reading_level(data)
            if 'final_math_level' not in data or not data['final_math_level']:
                data['final_math_level'] = compute_math_level(data)
            save_submission(data, request.remote_addr, session.get('fullname', ''))
            flash('✅ تم حفظ الاختبار بنجاح! يمكنك الآن إدخال حالة جديدة.')
            return redirect(url_for('assessment'))
        else:
            flash('❌ يرجى اختيار "نعم" للبدء أو "لا" للرفض.')
            return redirect(url_for('assessment'))
    return render_template_string(FORM_HTML, form=form, dist_by_gov=json.dumps(DIST_BY_GOV, ensure_ascii=False), school_by_dist=json.dumps(SCHOOL_BY_DIST, ensure_ascii=False))

if __name__ == '__main__':
    init_users()
    init_excel()
    app.run(debug=True, host='0.0.0.0', port=5000)